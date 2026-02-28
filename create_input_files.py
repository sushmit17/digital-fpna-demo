"""
Generate the 5 unit submission Excel files with realistic pharma data.
Each file represents one business unit's LBE submission for FY2026.
All figures in EUR millions.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
NAVY   = "1F3864"
BLUE   = "2E5FA3"
LGREY  = "F2F4F8"
MGREY  = "D6DCE4"
WHITE  = "FFFFFF"
GREEN  = "1A6B3C"
RED    = "A31A1A"

def hdr_font(white=True):
    return Font(name="Calibri", bold=True, size=11,
                color=WHITE if white else NAVY)

def body_font(bold=False, color="1A1A2E"):
    return Font(name="Calibri", bold=bold, size=10, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border():
    s = Side(style="thin", color="C5D0DC")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom():
    thin = Side(style="thin", color="C5D0DC")
    thick = Side(style="medium", color=NAVY)
    return Border(left=thin, right=thin, top=thin, bottom=thick)

def center():
    return Alignment(horizontal="center", vertical="center")

def right_align():
    return Alignment(horizontal="right", vertical="center")

def num_fmt(ws, cell_ref, value, is_total=False, is_negative=False):
    """Write a number with EUR millions format."""
    cell = ws[cell_ref]
    cell.value = value
    cell.number_format = '#,##0.0;(#,##0.0);"-"'
    cell.alignment = right_align()
    cell.border = border()
    if is_total:
        cell.font = body_font(bold=True, color=NAVY)
    elif is_negative:
        cell.font = body_font(color=RED)
    else:
        cell.font = body_font()

def make_unit_file(path, unit_name, unit_type, data):
    """
    data = dict of {line_name: {lbe, prior_lbe, budget, prior_year}}
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LBE FY2026"
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    ws["A1"] = f"{unit_name} — LBE FY2026 Submission"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    ws["A1"].fill = fill(NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center",
                                   indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Unit: {unit_name}  |  Type: {unit_type}  |  Currency: EUR millions"
    ws["A2"].font = Font(name="Calibri", size=9, color="5A6A80")
    ws["A2"].fill = fill("EEF2F8")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center",
                                   indent=1)
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 8  # spacer

    # ── Column headers ────────────────────────────────────────────────────────
    headers = ["P&L Line", "LBE FY2026", "Prior LBE", "Budget FY2026",
               "Prior Year (FY2025)"]
    for col, hdr in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=hdr)
        cell.font = hdr_font()
        cell.fill = fill(BLUE)
        cell.alignment = center() if col > 1 else Alignment(
            horizontal="left", vertical="center", indent=1)
        cell.border = border()
    ws.row_dimensions[4].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    totals = {"Net Sales", "Distribution Margin", "Total R&D",
              "Total SG&A", "Division Margin"}

    row = 5
    for line, vals in data.items():
        is_total = line in totals
        is_indent = not is_total

        # Label
        cell = ws.cell(row=row, column=1, value=line)
        cell.font = body_font(bold=is_total, color=NAVY if is_total else "1A1A2E")
        cell.fill = fill(LGREY if row % 2 == 0 else WHITE)
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   indent=1 if not is_total else 0)
        cell.border = thick_bottom() if is_total else border()
        ws.row_dimensions[row].height = 18

        for col_idx, key in enumerate(["lbe", "prior_lbe", "budget", "prior_year"], 2):
            c = ws.cell(row=row, column=col_idx)
            v = vals.get(key, 0)
            c.value = v
            c.number_format = '#,##0.0;(#,##0.0);"-"'
            c.alignment = right_align()
            c.fill = fill(LGREY if row % 2 == 0 else WHITE)
            c.font = body_font(bold=is_total, color=NAVY if is_total else "1A1A2E")
            c.border = thick_bottom() if is_total else border()

        row += 1

    # ── Footer note ───────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row+1}:E{row+1}")
    ws[f"A{row+1}"] = "Note: All figures in EUR millions. Positive = favourable. Submit to FP&A by deadline."
    ws[f"A{row+1}"].font = Font(name="Calibri", size=9, italic=True,
                                color="7A8FA8")
    ws[f"A{row+1}"].fill = fill("F8F9FB")

    wb.save(path)
    print(f"  Saved: {path}")


# ── DATA DEFINITIONS ─────────────────────────────────────────────────────────
# Realistic pharma commercial affiliate numbers (EUR millions)
# The variances tell a story: Immunology strong, Oncology soft vs budget,
# Neurology good vs PY. R&D slight overrun. Admin under budget.

IMMUNOLOGY_DATA = {
    "Net Sales":           {"lbe": 312.4, "prior_lbe": 305.1, "budget": 295.0, "prior_year": 278.6},
    "Distribution Margin": {"lbe": 249.9, "prior_lbe": 244.1, "budget": 236.0, "prior_year": 222.9},
}

ONCOLOGY_DATA = {
    "Net Sales":           {"lbe": 187.6, "prior_lbe": 196.2, "budget": 210.0, "prior_year": 172.4},
    "Distribution Margin": {"lbe": 140.7, "prior_lbe": 147.2, "budget": 157.5, "prior_year": 129.3},
}

NEUROLOGY_DATA = {
    "Net Sales":           {"lbe": 143.8, "prior_lbe": 141.3, "budget": 138.0, "prior_year": 122.5},
    "Distribution Margin": {"lbe": 107.9, "prior_lbe": 106.0, "budget": 103.5, "prior_year":  91.9},
}

RD_DATA = {
    "Total R&D":           {"lbe":  48.3, "prior_lbe":  46.1, "budget":  45.0, "prior_year":  41.2},
}

GENERAL_ADMIN_DATA = {
    "Marketing":                        {"lbe": 38.4, "prior_lbe": 39.1, "budget": 41.0, "prior_year": 35.8},
    "Other Advertising and Promotion":  {"lbe": 22.7, "prior_lbe": 23.4, "budget": 25.0, "prior_year": 20.1},
    "Sales Force":                      {"lbe": 54.2, "prior_lbe": 54.8, "budget": 56.0, "prior_year": 51.3},
    "General Admin":                    {"lbe": 19.8, "prior_lbe": 20.1, "budget": 21.5, "prior_year": 18.4},
    "Total SG&A":                       {"lbe": 135.1,"prior_lbe": 137.4,"budget": 143.5,"prior_year": 125.6},
}

base = "/home/claude/fpna_demo/input_files"
make_unit_file(f"{base}/immunology_LBE_FY2026.xlsx",
               "Immunology", "Sales Unit", IMMUNOLOGY_DATA)
make_unit_file(f"{base}/oncology_LBE_FY2026.xlsx",
               "Oncology", "Sales Unit", ONCOLOGY_DATA)
make_unit_file(f"{base}/neurology_LBE_FY2026.xlsx",
               "Neurology", "Sales Unit", NEUROLOGY_DATA)
make_unit_file(f"{base}/rd_LBE_FY2026.xlsx",
               "R&D", "R&D Unit", RD_DATA)
make_unit_file(f"{base}/general_admin_LBE_FY2026.xlsx",
               "General Admin", "Corporate Unit", GENERAL_ADMIN_DATA)

print("All 5 input files created.")
