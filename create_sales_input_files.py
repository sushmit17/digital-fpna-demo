"""
Rebuild the 3 Sales unit submission files with product-level detail.
Each file has:
  Tab 1 "Product Detail" — products with volumes (units 000s) and price (EUR)
  Tab 2 "LBE FY2026"    — P&L summary (Net Sales + Distribution Margin)
                           Net Sales is linked from the product detail tab.
All monetary figures in EUR millions. Volumes in thousands of units.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY   = "1F3864"
BLUE   = "2E5FA3"
TEAL   = "00B0A0"
LGREY  = "F2F4F8"
WHITE  = "FFFFFF"
GREEN  = "1A6B3C"
RED    = "A31A1A"
GOLD   = "7B5E00"

def thin_border():
    s = Side(style="thin", color="C5D0DC")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bot():
    t = Side(style="thin",   color="C5D0DC")
    b = Side(style="medium", color=NAVY)
    return Border(left=t, right=t, top=t, bottom=b)

def fill(c): return PatternFill("solid", fgColor=c)
def hfnt(c=WHITE): return Font(name="Calibri", bold=True, size=10, color=c)
def bfnt(bold=False, color="1A1A2E"): return Font(name="Calibri", bold=bold, size=10, color=color)
def aln(h="right"): return Alignment(horizontal=h, vertical="center")
def laln(): return Alignment(horizontal="left", vertical="center", indent=1)


def write_title(ws, text, sub, cols):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    ws["A1"] = text
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    ws["A1"].fill = fill(NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{get_column_letter(cols)}2")
    ws["A2"] = sub
    ws["A2"].font = Font(name="Calibri", size=9, color="5A6A80")
    ws["A2"].fill = fill("EEF2F8")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 8


def make_sales_unit(path, unit_name, products, dist_margin_pct):
    """
    products: list of dicts:
      {name, lbe_vol, plbe_vol, bgt_vol, py_vol,   <- volumes (000 units)
            lbe_px,  plbe_px,  bgt_px,  py_px}     <- avg net price (EUR/unit)
    dist_margin_pct: distribution margin as % of net sales
      {lbe, plbe, bgt, py}
    """
    wb = openpyxl.Workbook()

    # ── TAB 1: Product Detail ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Product Detail"
    ws1.sheet_view.showGridLines = False

    col_widths1 = [28, 10, 10, 10, 10, 10, 10, 10, 10, 12, 12, 12, 12]
    for i, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    write_title(ws1, f"{unit_name} — Product Detail  |  LBE FY2026",
                "Volumes in 000 units  |  Price = Average Net Price EUR/unit  |  Net Sales in EUR millions", 13)

    # Section headers
    # Row 4: blank spacer already set
    ws1.row_dimensions[4].height = 6

    # Volume header group — row 5
    ws1.merge_cells("B5:E5")
    ws1["B5"] = "Volume (000 units)"
    ws1["B5"].font = hfnt()
    ws1["B5"].fill = fill(BLUE)
    ws1["B5"].alignment = aln("center")
    for c in ["C5","D5","E5"]:
        ws1[c].fill = fill(BLUE)
        ws1[c].border = thin_border()
    ws1["B5"].border = thin_border()

    ws1.merge_cells("F5:I5")
    ws1["F5"] = "Avg Net Price (EUR/unit)"
    ws1["F5"].font = hfnt()
    ws1["F5"].fill = fill(TEAL)
    ws1["F5"].alignment = aln("center")
    for c in ["G5","H5","I5"]:
        ws1[c].fill = fill(TEAL)
        ws1[c].border = thin_border()
    ws1["F5"].border = thin_border()

    ws1.merge_cells("J5:M5")
    ws1["J5"] = "Net Sales (EUR millions)"
    ws1["J5"].font = hfnt()
    ws1["J5"].fill = fill(NAVY)
    ws1["J5"].alignment = aln("center")
    for c in ["K5","L5","M5"]:
        ws1[c].fill = fill(NAVY)
        ws1[c].border = thin_border()
    ws1["J5"].border = thin_border()
    ws1.row_dimensions[5].height = 20

    # Sub-headers — row 6
    sub_hdrs = ["Product", "LBE", "Prior LBE", "Budget", "Prior Year",
                "LBE", "Prior LBE", "Budget", "Prior Year",
                "LBE", "Prior LBE", "Budget", "Prior Year"]
    sub_fills = [NAVY] + [BLUE]*4 + [TEAL]*4 + [NAVY]*4
    for ci, (h, f_col) in enumerate(zip(sub_hdrs, sub_fills), 1):
        c = ws1.cell(row=6, column=ci, value=h)
        c.font = hfnt()
        c.fill = fill(f_col)
        c.alignment = laln() if ci == 1 else aln("center")
        c.border = thin_border()
    ws1.row_dimensions[6].height = 18

    # Data rows — starting row 7
    data_rows = []
    for pi, prod in enumerate(products):
        row = 7 + pi
        alt = "F2F4F8" if pi % 2 == 0 else WHITE

        # Product name
        c = ws1.cell(row=row, column=1, value=prod["name"])
        c.font = bfnt(bold=True, color=NAVY)
        c.fill = fill(alt)
        c.alignment = laln()
        c.border = thin_border()

        # Volumes (B–E)
        for ci, key in enumerate(["lbe_vol","plbe_vol","bgt_vol","py_vol"], 2):
            cell = ws1.cell(row=row, column=ci, value=prod[key])
            cell.number_format = '#,##0;(#,##0);"-"'
            cell.alignment = aln()
            cell.fill = fill(alt)
            cell.border = thin_border()
            cell.font = bfnt()

        # Prices (F–I)
        for ci, key in enumerate(["lbe_px","plbe_px","bgt_px","py_px"], 6):
            cell = ws1.cell(row=row, column=ci, value=prod[key])
            cell.number_format = '#,##0.00;(#,##0.00);"-"'
            cell.alignment = aln()
            cell.fill = fill(alt)
            cell.border = thin_border()
            cell.font = bfnt()

        # Net Sales formulas (J–M) = Vol * Price / 1000 (to convert to EUR millions)
        for ci, (vol_col, px_col) in enumerate(
                [("B","F"),("C","G"),("D","H"),("E","I")], 10):
            col_letter = get_column_letter(ci)
            formula = f"=ROUND({vol_col}{row}*{px_col}{row}/1000000,1)"
            cell = ws1.cell(row=row, column=ci, value=formula)
            cell.number_format = '#,##0.0;(#,##0.0);"-"'
            cell.alignment = aln()
            cell.fill = fill(alt)
            cell.border = thin_border()
            cell.font = bfnt(color=GREEN)  # green = formula

        data_rows.append(row)
        ws1.row_dimensions[row].height = 18

    # Total row
    total_row = 7 + len(products)
    ws1.cell(row=total_row, column=1, value="Total Net Sales").font = bfnt(bold=True, color=NAVY)
    ws1.cell(row=total_row, column=1).fill = fill("EEF2F8")
    ws1.cell(row=total_row, column=1).alignment = laln()
    ws1.cell(row=total_row, column=1).border = thick_bot()

    for ci in range(2, 10):  # blank volume/price totals
        c = ws1.cell(row=total_row, column=ci, value="")
        c.fill = fill("EEF2F8")
        c.border = thick_bot()

    first_data = 7
    last_data  = 6 + len(products)
    for ci_i, ci in enumerate(range(10, 14)):
        col_l = get_column_letter(ci)
        formula = f"=SUM({col_l}{first_data}:{col_l}{last_data})"
        c = ws1.cell(row=total_row, column=ci, value=formula)
        c.number_format = '#,##0.0;(#,##0.0);"-"'
        c.alignment = aln()
        c.fill = fill("EEF2F8")
        c.border = thick_bot()
        c.font = bfnt(bold=True, color=NAVY)

    ws1.row_dimensions[total_row].height = 20
    ws1.freeze_panes = "B7"

    # ── TAB 2: LBE FY2026 (P&L Summary) ──────────────────────────────────────
    ws2 = wb.create_sheet("LBE FY2026")
    ws2.sheet_view.showGridLines = False

    col_widths2 = [32, 14, 14, 14, 14]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    write_title(ws2,
                f"{unit_name} — LBE FY2026 Submission",
                f"Unit: {unit_name}  |  Type: Sales Unit  |  Currency: EUR millions", 5)

    # Column headers — row 4
    hdrs = ["P&L Line", "LBE FY2026", "Prior LBE", "Budget FY2026", "Prior Year (FY2025)"]
    for ci, h in enumerate(hdrs, 1):
        c = ws2.cell(row=4, column=ci, value=h)
        c.font = hfnt()
        c.fill = fill(BLUE)
        c.alignment = laln() if ci == 1 else aln("center")
        c.border = thin_border()
    ws2.row_dimensions[4].height = 20

    ns_row = total_row  # row in Product Detail sheet where total net sales lives
    # Net Sales — pull from Product Detail via formula
    ns_cols = {
        "LBE FY2026":        f"='Product Detail'!J{ns_row}",
        "Prior LBE":         f"='Product Detail'!K{ns_row}",
        "Budget FY2026":     f"='Product Detail'!L{ns_row}",
        "Prior Year (FY2025)": f"='Product Detail'!M{ns_row}",
    }

    # Distribution margin = Net Sales * margin %
    dm_pct = dist_margin_pct

    pnl_lines = [
        ("Net Sales", ns_cols, False),
        ("Distribution Margin", {
            "LBE FY2026":          f"=ROUND(B5*{dm_pct['lbe']},1)",
            "Prior LBE":           f"=ROUND(C5*{dm_pct['plbe']},1)",
            "Budget FY2026":       f"=ROUND(D5*{dm_pct['bgt']},1)",
            "Prior Year (FY2025)": f"=ROUND(E5*{dm_pct['py']},1)",
        }, False),
    ]

    for ri, (line, vals, _) in enumerate(pnl_lines):
        row = 5 + ri
        is_total = True
        alt = "EEF2F8" if is_total else (LGREY if row % 2 == 0 else WHITE)
        c = ws2.cell(row=row, column=1, value=line)
        c.font = bfnt(bold=True, color=NAVY)
        c.fill = fill(alt)
        c.alignment = laln()
        c.border = thick_bot()
        ws2.row_dimensions[row].height = 20

        for ci, col_name in enumerate(["LBE FY2026","Prior LBE","Budget FY2026","Prior Year (FY2025)"], 2):
            cell = ws2.cell(row=row, column=ci, value=vals[col_name])
            cell.number_format = '#,##0.0;(#,##0.0);"-"'
            cell.alignment = aln()
            cell.fill = fill(alt)
            cell.border = thick_bot()
            cell.font = bfnt(bold=True, color=GREEN)  # formula

    # Footer
    note_row = 5 + len(pnl_lines) + 1
    ws2.merge_cells(f"A{note_row}:E{note_row}")
    ws2[f"A{note_row}"] = (
        "Note: Net Sales linked from Product Detail tab (Vol × Price / 1,000,000). "
        "Distribution Margin = Net Sales × margin %. All figures EUR millions."
    )
    ws2[f"A{note_row}"].font = Font(name="Calibri", size=9, italic=True, color="7A8FA8")
    ws2[f"A{note_row}"].fill = fill("F8F9FB")

    wb.save(path)
    print(f"  Saved: {path}")


# ── IMMUNOLOGY ─────────────────────────────────────────────────────────────────
# Products: Immuno-A (flagship biologic), Immuno-B (new launch), Immuno-C (legacy)
make_sales_unit(
    "input_files/immunology_LBE_FY2026.xlsx",
    "Immunology",
    products=[
        {"name": "Immuno-A (Biologic, IV)",
         "lbe_vol": 42800, "plbe_vol": 41500, "bgt_vol": 40000, "py_vol": 36200,
         "lbe_px": 4850, "plbe_px": 4830, "bgt_px": 4800, "py_px": 4750},
        {"name": "Immuno-B (SC, New Launch)",
         "lbe_vol": 18600, "plbe_vol": 17200, "bgt_vol": 16000, "py_vol":  8400,
         "lbe_px": 3920, "plbe_px": 3900, "bgt_px": 3850, "py_px": 3800},
        {"name": "Immuno-C (Legacy Oral)",
         "lbe_vol": 31400, "plbe_vol": 32100, "bgt_vol": 33000, "py_vol": 38800,
         "lbe_px": 1680, "plbe_px": 1690, "bgt_px": 1700, "py_px": 1720},
    ],
    dist_margin_pct={"lbe": 0.800, "plbe": 0.800, "bgt": 0.800, "py": 0.800},
)

# ── ONCOLOGY ───────────────────────────────────────────────────────────────────
# Products: Onco-Prime (1L solid tumours), Onco-Next (2L, competitive pressure)
make_sales_unit(
    "input_files/oncology_LBE_FY2026.xlsx",
    "Oncology",
    products=[
        {"name": "Onco-Prime (1L Solid Tumours)",
         "lbe_vol": 12400, "plbe_vol": 13200, "bgt_vol": 14800, "py_vol": 11600,
         "lbe_px": 8950, "plbe_px": 8950, "bgt_px": 9000, "py_px": 8800},
        {"name": "Onco-Next (2L, Haematology)",
         "lbe_vol":  8200, "plbe_vol":  8600, "bgt_vol":  9200, "py_vol":  6900,
         "lbe_px": 7640, "plbe_px": 7620, "bgt_px": 7600, "py_px": 7500},
        {"name": "Onco-Support (Supportive Care)",
         "lbe_vol": 22100, "plbe_vol": 22400, "bgt_vol": 22000, "py_vol": 20800,
         "lbe_px":  1340, "plbe_px":  1350, "bgt_px":  1350, "py_px":  1320},
    ],
    dist_margin_pct={"lbe": 0.750, "plbe": 0.750, "bgt": 0.750, "py": 0.750},
)

# ── NEUROLOGY ──────────────────────────────────────────────────────────────────
# Products: Neuro-Alpha (MS, established), Neuro-Beta (epilepsy, growing)
make_sales_unit(
    "input_files/neurology_LBE_FY2026.xlsx",
    "Neurology",
    products=[
        {"name": "Neuro-Alpha (MS, Established)",
         "lbe_vol": 28600, "plbe_vol": 28100, "bgt_vol": 27500, "py_vol": 24800,
         "lbe_px": 3420, "plbe_px": 3410, "bgt_px": 3400, "py_px": 3350},
        {"name": "Neuro-Beta (Epilepsy, SC)",
         "lbe_vol": 19800, "plbe_vol": 19400, "bgt_vol": 19000, "py_vol": 15600,
         "lbe_px": 2180, "plbe_px": 2175, "bgt_px": 2150, "py_px": 2100},
        {"name": "Neuro-Gamma (Pain, OTC Bridge)",
         "lbe_vol": 41200, "plbe_vol": 41500, "bgt_vol": 40000, "py_vol": 38400,
         "lbe_px":   620, "plbe_px":   618, "bgt_px":   615, "py_px":   605},
    ],
    dist_margin_pct={"lbe": 0.750, "plbe": 0.750, "bgt": 0.750, "py": 0.750},
)

print("\nAll 3 sales unit files rebuilt with product detail.")
